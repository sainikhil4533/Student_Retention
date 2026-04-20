from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, TextIOWrapper
import csv
import re
import time
import zipfile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy.orm import Session
import openpyxl

from src.api.auth import AuthContext, require_roles
from src.api.auth_accounts import provision_counsellor_auth_accounts, provision_student_auth_account
from src.api.dependencies import prediction_service
from src.api.scoring_service import score_student_from_db
from src.api.schemas import InstitutionImportResponse, VignanImportResponse
from src.db.database import get_db
from src.db.repository import EventRepository
from src.worker.job_queue import (
    enqueue_faculty_alert_email_job,
    enqueue_student_warning_email_job,
)


router = APIRouter(prefix="/admin/imports", tags=["admin-imports"])

REQUIRED_SHEETS = {
    "Admissions",
    "Academics",
    "Attendance",
    "Finance",
    "Registration",
    "LMS",
}

OPTIONAL_SHEETS = {"SupportMapping"}

INSTITUTION_REQUIRED_SHEETS = {
    "Admissions",
    "Registration",
    "Attendance",
    "AttendancePolicy",
    "SubjectCatalog",
}

INSTITUTION_OPTIONAL_SHEETS = {
    "SupportMapping",
    "StudentAcademicProgress",
    "SemesterProgress",
    "Academics",
    "Finance",
    "LMS",
}

REQUIRED_COLUMNS = {
    "Admissions": {
        "registerno",
        "Branch",
        "Batch",
        "Category",
        "Region",
        "ParentEdu",
        "Occupation",
        "Income",
        "Gender",
        "AgeBand",
        "Attempts",
    },
    "Academics": {"registerno", "Semester", "CGPA", "Backlogs", "shortname", "Marks"},
    "Attendance": {
        "registerno",
        "Semester",
        "Overall%",
        "shortname",
        "Subject%",
        "ConsecutiveAbs",
        "MissedvDays",
        "Trend",
    },
    "Finance": {"registerno", "Status", "Due", "DelayDays", "Scholarship", "Modifier"},
    "Registration": {"registerno", "Semester", "Registered", "FinalStatus"},
    "LMS": {
        "registerno",
        "event_date",
        "id_site",
        "sum_click",
        "engagement_tag",
        "resource_type",
    },
}

OPTIONAL_REQUIRED_COLUMNS = {
    "SupportMapping": {
        "registerno",
        "student_email",
        "disability_status",
        "faculty_name",
        "faculty_email",
        "parent_name",
        "parent_relationship",
        "parent_email",
        "parent_phone",
        "preferred_guardian_channel",
        "guardian_contact_enabled",
        "counsellor_name",
        "counsellor_email",
    }
}

FAIRNESS_REVIEW_REQUIRED_PROFILE_FIELDS = (
    "category",
    "region",
    "income",
    "parent_education",
    "occupation",
)

ATTENDANCE_COLUMN_ALIASES = {
    "Overall%": {"Overall%", "OverallPer", "OverallPercent", "OverallAttendancePercent"},
    "Subject%": {"Subject%", "SubjectPer", "SubjectPercent", "SubjectAttendancePercent"},
    "SubjectCode": {"SubjectCode", "SubjectCo"},
    "shortname": {"shortname", "SubjectName", "Subject"},
}


@dataclass
class ImportCounters:
    profiles_upserted: int = 0
    support_mappings_applied: int = 0
    accounts_provisioned: int = 0
    attendance_policies_upserted: int = 0
    subject_catalog_rows_upserted: int = 0
    academic_progress_rows_upserted: int = 0
    semester_progress_rows_upserted: int = 0
    attendance_rows_upserted: int = 0
    academic_rows_upserted: int = 0
    lms_events_ingested: int = 0
    erp_events_ingested: int = 0
    finance_events_ingested: int = 0
    scoring_triggered: int = 0


@router.post("/vignan", response_model=VignanImportResponse)
async def import_vignan(
    file: UploadFile = File(...),
    trigger_scoring: bool = True,
    dry_run: bool = False,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_roles("admin")),
) -> VignanImportResponse:
    if file.filename is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload a .xlsx or .zip file.",
        )

    filename = file.filename.lower()
    payload = await file.read()
    if filename.endswith(".xlsx"):
        sheets = _read_excel(BytesIO(payload))
    elif filename.endswith(".zip"):
        sheets = _read_zip_csv(BytesIO(payload))
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .xlsx or a .zip of CSV sheets.",
        )

    missing = REQUIRED_SHEETS - set(sheets.keys())
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing sheets: {', '.join(sorted(missing))}.",
        )

    missing_columns = _validate_required_columns(sheets)
    if missing_columns:
        formatted = "; ".join(
            f"{sheet}: {', '.join(sorted(cols))}"
            for sheet, cols in missing_columns.items()
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns -> {formatted}.",
        )

    repository = EventRepository(db)
    counters = ImportCounters()
    errors: list[str] = []

    admissions = _group_by_registerno(sheets["Admissions"])
    academics = _group_by_registerno(sheets["Academics"])
    attendance = _group_by_registerno(sheets["Attendance"])
    finance = _group_by_registerno(sheets["Finance"])
    registration = _group_by_registerno(sheets["Registration"])
    lms = _group_by_registerno(sheets["LMS"])
    support_mapping = _group_by_registerno(sheets.get("SupportMapping", []))

    all_regs = set(admissions) | set(academics) | set(attendance) | set(finance) | set(
        registration
    ) | set(lms)

    next_student_id = _next_student_id(repository)
    for reg in sorted(all_regs):
        try:
            profile = repository.get_student_profile_by_external_ref(reg)
            if profile is None:
                student_id = next_student_id
                next_student_id += 1
            else:
                student_id = int(profile.student_id)

            admissions_row = _first_row(admissions.get(reg))
            registration_row = _first_row(registration.get(reg))
            support_mapping_row = _first_row(support_mapping.get(reg))
            profile_payload = _build_profile_payload(
                student_id=student_id,
                registerno=reg,
                admissions_row=admissions_row,
                registration_row=registration_row,
                support_mapping_row=support_mapping_row,
            )
            if support_mapping_row is not None:
                counters.support_mappings_applied += 1
            if not dry_run:
                persisted_profile = repository.upsert_student_profile(profile_payload)
                _, account_created = provision_student_auth_account(
                    repository,
                    profile=persisted_profile,
                )
                if account_created:
                    counters.accounts_provisioned += 1
            counters.profiles_upserted += 1

            for lms_row in lms.get(reg, []):
                event = _build_lms_event(student_id, reg, lms_row)
                if not dry_run:
                    repository.add_lms_event(event)
                counters.lms_events_ingested += 1

            erp_event = _build_erp_event(
                student_id=student_id,
                registerno=reg,
                academics_rows=academics.get(reg, []),
                attendance_rows=attendance.get(reg, []),
                admissions_row=admissions_row,
                registration_row=registration_row,
            )
            if erp_event is not None:
                if not dry_run:
                    repository.add_erp_event(erp_event)
                counters.erp_events_ingested += 1

            finance_row = _first_row(finance.get(reg))
            if finance_row:
                finance_event = _build_finance_event(student_id, reg, finance_row)
                if not dry_run:
                    repository.add_finance_event(finance_event)
                counters.finance_events_ingested += 1

            if (
                trigger_scoring
                and not dry_run
                and erp_event is not None
                and lms.get(reg)
            ):
                score_result = score_student_from_db(
                    student_id=student_id,
                    db=db,
                    prediction_service=prediction_service,
                )
                if (
                    score_result.get("student_warning_triggered")
                    and score_result.get("student_warning_status") == "pending"
                    and score_result.get("student_warning_event_id") is not None
                ):
                    enqueue_student_warning_email_job(
                        warning_event_id=int(score_result["student_warning_event_id"]),
                        student_id=student_id,
                        prediction_history_id=int(score_result["prediction_history_id"]),
                        warning_type=str(score_result["student_warning_type"]),
                        recipient=str(profile_payload.get("student_email") or "unconfigured"),
                    )
                if (
                    score_result.get("alert_triggered")
                    and score_result.get("alert_status") == "pending"
                    and score_result.get("alert_event_id") is not None
                ):
                    enqueue_faculty_alert_email_job(
                        alert_event_id=int(score_result["alert_event_id"]),
                        student_id=student_id,
                        prediction_history_id=int(score_result["prediction_history_id"]),
                        alert_type=str(score_result["alert_type"]),
                    )
                counters.scoring_triggered += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{reg}: {exc}")

    return VignanImportResponse(
        status="completed",
        total_students=len(all_regs),
        profiles_upserted=counters.profiles_upserted,
        support_mappings_applied=counters.support_mappings_applied,
        accounts_provisioned=counters.accounts_provisioned,
        lms_events_ingested=counters.lms_events_ingested,
        erp_events_ingested=counters.erp_events_ingested,
        finance_events_ingested=counters.finance_events_ingested,
        scoring_triggered=counters.scoring_triggered,
        errors=errors,
    )


@router.post("/institution", response_model=InstitutionImportResponse)
async def import_institution(
    file: UploadFile = File(...),
    trigger_scoring: bool = True,
    dry_run: bool = False,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(require_roles("admin")),
) -> InstitutionImportResponse:
    del auth
    started_at = time.perf_counter()
    if file.filename is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Upload a .xlsx or .zip file.",
        )

    filename = file.filename.lower()
    print(f"[institution.import] start filename={filename} trigger_scoring={trigger_scoring} dry_run={dry_run}")
    payload = await file.read()
    print(f"[institution.import] file read complete bytes={len(payload)} elapsed={time.perf_counter() - started_at:.2f}s")
    if filename.endswith(".xlsx"):
        sheets = _read_excel(BytesIO(payload))
    elif filename.endswith(".zip"):
        sheets = _read_zip_csv(BytesIO(payload))
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use .xlsx or a .zip of CSV sheets.",
        )
    print(f"[institution.import] workbook parsed sheets={sorted(sheets.keys())} elapsed={time.perf_counter() - started_at:.2f}s")

    missing = INSTITUTION_REQUIRED_SHEETS - set(sheets.keys())
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing sheets: {', '.join(sorted(missing))}.",
        )

    missing_columns = _validate_institution_contract_columns(sheets)
    if missing_columns:
        formatted = "; ".join(
            f"{sheet}: {', '.join(sorted(cols))}"
            for sheet, cols in missing_columns.items()
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns -> {formatted}.",
        )
    print(f"[institution.import] contract validation complete elapsed={time.perf_counter() - started_at:.2f}s")

    repository = EventRepository(db)
    counters = ImportCounters()
    errors: list[str] = []

    admissions = _group_by_registerno(sheets["Admissions"])
    registration = _group_by_registerno(sheets["Registration"])
    attendance = _group_by_registerno(sheets["Attendance"])
    support_mapping = _group_by_registerno(sheets.get("SupportMapping", []))
    academic_progress = _group_by_registerno(sheets.get("StudentAcademicProgress", []))
    semester_progress = _group_by_registerno(sheets.get("SemesterProgress", []))
    academics = _group_by_registerno(sheets.get("Academics", []))
    finance = _group_by_registerno(sheets.get("Finance", []))
    lms = _group_by_registerno(sheets.get("LMS", []))

    institution_name = _detect_institution_name(sheets)
    policy_records = [
        _build_policy_record(row)
        for row in sheets.get("AttendancePolicy", [])
        if _safe_str(row.get("InstitutionName"))
    ]
    if not policy_records:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="AttendancePolicy must contain at least one valid row.",
        )
    policy = policy_records[0]

    subject_catalog_records = [
        _build_subject_catalog_record(row)
        for row in sheets.get("SubjectCatalog", [])
        if _safe_str(row.get("SubjectCode")) and _safe_str(row.get("SubjectName"))
    ]
    if not subject_catalog_records:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="SubjectCatalog must contain at least one valid row.",
        )
    print(
        "[institution.import] prepared policy and catalog "
        f"policy_rows={len(policy_records)} subject_catalog_rows={len(subject_catalog_records)} "
        f"elapsed={time.perf_counter() - started_at:.2f}s"
    )

    if not dry_run:
        repository.replace_institution_attendance_policies(
            institution_name=institution_name,
            records=policy_records,
            commit=False,
        )
        repository.replace_subject_catalog_entries(
            institution_name=institution_name,
            records=subject_catalog_records,
            commit=False,
        )
        db.commit()
    counters.attendance_policies_upserted = len(policy_records)
    counters.subject_catalog_rows_upserted = len(subject_catalog_records)

    subject_catalog_index = _build_subject_catalog_index(subject_catalog_records)

    all_regs = (
        set(admissions)
        | set(registration)
        | set(attendance)
        | set(support_mapping)
        | set(academic_progress)
        | set(semester_progress)
        | set(academics)
        | set(finance)
        | set(lms)
    )
    print(f"[institution.import] grouped rows total_students={len(all_regs)} elapsed={time.perf_counter() - started_at:.2f}s")

    next_student_id = _next_student_id(repository)
    for index, reg in enumerate(sorted(all_regs), start=1):
        try:
            student_has_pending_writes = False
            existing_profile = repository.get_student_profile_by_external_ref(reg)
            if existing_profile is None:
                student_id = next_student_id
                next_student_id += 1
            else:
                student_id = int(existing_profile.student_id)

            admissions_row = _first_row(admissions.get(reg))
            registration_row = _first_row(registration.get(reg))
            support_mapping_row = _first_row(support_mapping.get(reg))
            academic_progress_row = _first_row(academic_progress.get(reg))

            profile_payload = _build_profile_payload_generalized(
                student_id=student_id,
                registerno=reg,
                admissions_row=admissions_row,
                registration_row=registration_row,
                support_mapping_row=support_mapping_row,
                academic_progress_row=academic_progress_row,
                institution_name=institution_name,
            )
            if support_mapping_row is not None:
                counters.support_mappings_applied += 1
            if not dry_run:
                persisted_profile = repository.upsert_student_profile(profile_payload, commit=False)
                _, account_created = provision_student_auth_account(
                    repository,
                    profile=persisted_profile,
                    commit=False,
                )
                if account_created:
                    counters.accounts_provisioned += 1
                student_has_pending_writes = True
            counters.profiles_upserted += 1

            progress_record = _build_student_academic_progress_record(
                student_id=student_id,
                registerno=reg,
                institution_name=institution_name,
                admissions_row=admissions_row,
                registration_row=registration_row,
                academic_progress_row=academic_progress_row,
            )
            if progress_record is not None:
                if not dry_run:
                    repository.replace_student_academic_progress(
                        student_id=student_id,
                        record=progress_record,
                        commit=False,
                    )
                    student_has_pending_writes = True
                counters.academic_progress_rows_upserted += 1

            attendance_rows = attendance.get(reg, [])
            built_attendance_rows = _build_subject_attendance_records(
                student_id=student_id,
                registerno=reg,
                institution_name=institution_name,
                admissions_row=admissions_row,
                attendance_rows=attendance_rows,
                policy=policy,
                subject_catalog_index=subject_catalog_index,
            )
            if built_attendance_rows:
                if not dry_run:
                    repository.replace_student_subject_attendance(
                        student_id=student_id,
                        records=built_attendance_rows,
                        commit=False,
                    )
                    student_has_pending_writes = True
                counters.attendance_rows_upserted += len(built_attendance_rows)

            semester_progress_rows = _build_semester_progress_records(
                student_id=student_id,
                registerno=reg,
                provided_rows=semester_progress.get(reg, []),
                attendance_rows=built_attendance_rows,
                registration_row=registration_row,
                academic_progress_row=academic_progress_row,
                policy=policy,
            )
            if semester_progress_rows:
                if not dry_run:
                    repository.replace_student_semester_progress(
                        student_id=student_id,
                        records=semester_progress_rows,
                        commit=False,
                    )
                    student_has_pending_writes = True
                counters.semester_progress_rows_upserted += len(semester_progress_rows)

            academic_rows = _build_student_academic_records(
                student_id=student_id,
                registerno=reg,
                institution_name=institution_name,
                admissions_row=admissions_row,
                source_rows=academics.get(reg, []),
                subject_catalog_index=subject_catalog_index,
                attendance_rows=built_attendance_rows,
            )
            if academic_rows:
                if not dry_run:
                    repository.replace_student_academic_records(
                        student_id=student_id,
                        records=academic_rows,
                        commit=False,
                    )
                    student_has_pending_writes = True
                counters.academic_rows_upserted += len(academic_rows)

            for lms_row in lms.get(reg, []):
                event = _build_lms_event_generic(student_id, reg, lms_row)
                if not dry_run:
                    repository.add_lms_event(event, commit=False)
                    student_has_pending_writes = True
                counters.lms_events_ingested += 1

            finance_row = _first_row(finance.get(reg))
            if finance_row:
                finance_event = _build_finance_event_generic(student_id, reg, finance_row)
                if not dry_run:
                    repository.add_finance_event(finance_event, commit=False)
                    student_has_pending_writes = True
                counters.finance_events_ingested += 1

            erp_event = _build_erp_event_generalized(
                student_id=student_id,
                registerno=reg,
                institution_name=institution_name,
                admissions_row=admissions_row,
                registration_row=registration_row,
                attendance_rows=built_attendance_rows,
                academic_rows=academic_rows,
            )
            if erp_event is not None:
                if not dry_run:
                    repository.add_erp_event(erp_event, commit=False)
                    student_has_pending_writes = True
                counters.erp_events_ingested += 1

            if (
                trigger_scoring
                and not dry_run
                and erp_event is not None
                and lms.get(reg)
            ):
                if student_has_pending_writes:
                    db.commit()
                    student_has_pending_writes = False
                score_result = score_student_from_db(
                    student_id=student_id,
                    db=db,
                    prediction_service=prediction_service,
                )
                if (
                    score_result.get("student_warning_triggered")
                    and score_result.get("student_warning_status") == "pending"
                    and score_result.get("student_warning_event_id") is not None
                ):
                    enqueue_student_warning_email_job(
                        warning_event_id=int(score_result["student_warning_event_id"]),
                        student_id=student_id,
                        prediction_history_id=int(score_result["prediction_history_id"]),
                        warning_type=str(score_result["student_warning_type"]),
                        recipient=str(profile_payload.get("student_email") or "unconfigured"),
                    )
                if (
                    score_result.get("alert_triggered")
                    and score_result.get("alert_status") == "pending"
                    and score_result.get("alert_event_id") is not None
                ):
                    enqueue_faculty_alert_email_job(
                        alert_event_id=int(score_result["alert_event_id"]),
                        student_id=student_id,
                        prediction_history_id=int(score_result["prediction_history_id"]),
                        alert_type=str(score_result["alert_type"]),
                    )
                counters.scoring_triggered += 1
            elif not dry_run and student_has_pending_writes:
                db.commit()
        except Exception as exc:  # noqa: BLE001
            if not dry_run:
                db.rollback()
            errors.append(f"{reg}: {exc}")
        if index % 25 == 0 or index == len(all_regs):
            print(
                "[institution.import] progress "
                f"processed={index}/{len(all_regs)} attendance_rows={counters.attendance_rows_upserted} "
                f"academic_rows={counters.academic_rows_upserted} elapsed={time.perf_counter() - started_at:.2f}s"
            )

    if not dry_run:
        imported_profiles = repository.get_imported_student_profiles()
        counsellor_accounts_created = provision_counsellor_auth_accounts(
            repository,
            profiles=imported_profiles,
            commit=True,
        )
        counters.accounts_provisioned += counsellor_accounts_created
        print(
            "[institution.import] counsellor accounts provisioned "
            f"created={counsellor_accounts_created} total_accounts={counters.accounts_provisioned} "
            f"elapsed={time.perf_counter() - started_at:.2f}s"
        )

    print(
        "[institution.import] completed "
        f"students={len(all_regs)} profiles={counters.profiles_upserted} "
        f"attendance={counters.attendance_rows_upserted} academics={counters.academic_rows_upserted} "
        f"errors={len(errors)} elapsed={time.perf_counter() - started_at:.2f}s"
    )
    return InstitutionImportResponse(
        status="completed",
        import_type="institution",
        institution_name=institution_name,
        total_students=len(all_regs),
        profiles_upserted=counters.profiles_upserted,
        support_mappings_applied=counters.support_mappings_applied,
        accounts_provisioned=counters.accounts_provisioned,
        attendance_policies_upserted=counters.attendance_policies_upserted,
        subject_catalog_rows_upserted=counters.subject_catalog_rows_upserted,
        academic_progress_rows_upserted=counters.academic_progress_rows_upserted,
        semester_progress_rows_upserted=counters.semester_progress_rows_upserted,
        attendance_rows_upserted=counters.attendance_rows_upserted,
        academic_rows_upserted=counters.academic_rows_upserted,
        lms_events_ingested=counters.lms_events_ingested,
        erp_events_ingested=counters.erp_events_ingested,
        finance_events_ingested=counters.finance_events_ingested,
        scoring_triggered=counters.scoring_triggered,
        errors=errors,
    )


def _read_excel(buffer: BytesIO) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
    sheets: dict[str, list[dict]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        missing_headers = [header for header in headers if header in (None, "")]
        if missing_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Sheet {name} has empty header cells.",
            )
        rows: list[dict] = []
        for r in range(2, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(value in (None, "") for value in values):
                continue
            row = {str(headers[i]).strip(): values[i] for i in range(len(headers))}
            rows.append(row)
        sheets[str(name).strip()] = rows
    return sheets


def _read_zip_csv(buffer: BytesIO) -> dict[str, list[dict]]:
    sheets: dict[str, list[dict]] = {}
    with zipfile.ZipFile(buffer) as archive:
        for info in archive.infolist():
            name = info.filename.split("/")[-1]
            if not name.lower().endswith(".csv"):
                continue
            sheet_name = name[:-4]
            with archive.open(info) as handle:
                wrapper = TextIOWrapper(handle, encoding="utf-8")
                reader = csv.DictReader(wrapper)
                if not reader.fieldnames or any(
                    name in (None, "") for name in reader.fieldnames
                ):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Sheet {sheet_name} has empty header cells.",
                    )
                sheets[sheet_name] = [row for row in reader if any(row.values())]
    return sheets


def _validate_required_columns(sheets: dict[str, list[dict]]) -> dict[str, set[str]]:
    missing: dict[str, set[str]] = {}
    for sheet, required in REQUIRED_COLUMNS.items():
        rows = sheets.get(sheet, [])
        if not rows:
            continue
        present = set(str(key).strip() for key in rows[0].keys())
        missing_columns = set(required) - present
        if missing_columns:
            missing[sheet] = missing_columns
    for sheet, required in OPTIONAL_REQUIRED_COLUMNS.items():
        rows = sheets.get(sheet, [])
        if not rows:
            continue
        present = set(str(key).strip() for key in rows[0].keys())
        missing_columns = set(required) - present
        if missing_columns:
            missing[sheet] = missing_columns
    return missing


def _validate_institution_contract_columns(sheets: dict[str, list[dict]]) -> dict[str, set[str]]:
    missing: dict[str, set[str]] = {}
    base_required = {
        "Admissions": {"registerno", "Branch", "Batch", "Gender", "AgeBand", "Attempts"},
        "Registration": {"registerno", "Semester", "FinalStatus"},
        "AttendancePolicy": {
            "InstitutionName",
            "OverallMinPercent",
            "SubjectMinPercent",
            "RGradeBelowPercent",
            "IGradeMinPercent",
            "IGradeMaxPercent",
        },
        "SubjectCatalog": {
            "InstitutionName",
            "Branch",
            "Year",
            "Semester",
            "SubjectCode",
            "SubjectName",
        },
        "Attendance": {"registerno", "Semester", "Overall%", "Subject%"},
    }
    for sheet, required in base_required.items():
        rows = sheets.get(sheet, [])
        if not rows:
            continue
        present = {str(key).strip() for key in rows[0].keys()}
        missing_columns: set[str] = set()
        for column_name in required:
            aliases = ATTENDANCE_COLUMN_ALIASES.get(column_name, {column_name}) if sheet == "Attendance" else {column_name}
            if not any(alias in present for alias in aliases):
                missing_columns.add(column_name)
        if sheet == "Attendance" and not any(
            alias in present for alias in (ATTENDANCE_COLUMN_ALIASES["SubjectCode"] | ATTENDANCE_COLUMN_ALIASES["shortname"])
        ):
            missing_columns.add("SubjectCode or shortname")
        if missing_columns:
            missing[sheet] = missing_columns
    optional_support_rows = sheets.get("SupportMapping", [])
    if optional_support_rows:
        present = {str(key).strip() for key in optional_support_rows[0].keys()}
        optional_required = {
            "registerno",
            "student_email",
            "faculty_name",
            "faculty_email",
            "counsellor_name",
            "counsellor_email",
        }
        missing_columns = optional_required - present
        if missing_columns:
            missing["SupportMapping"] = missing_columns
    return missing


def _group_by_registerno(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        reg = _normalize_registerno(row.get("registerno"))
        if reg:
            grouped[reg].append(row)
    return grouped


def _normalize_registerno(value) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _next_student_id(repository: EventRepository) -> int:
    base = repository.get_max_student_id()
    return max(base, 880000) + 1


def _first_row(rows: list[dict] | None) -> dict | None:
    if not rows:
        return None
    return rows[0]


def _row_value(row: dict | None, *keys: str):
    if not row:
        return None
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return None


def _build_profile_payload(
    *,
    student_id: int,
    registerno: str,
    admissions_row: dict | None,
    registration_row: dict | None,
    support_mapping_row: dict | None,
) -> dict:
    admissions = admissions_row or {}
    registration = registration_row or {}
    support_mapping = support_mapping_row or {}
    gender = _safe_str(admissions.get("Gender")) or "U"
    age_band = _safe_str(admissions.get("AgeBand")) or "Unknown"
    attempts = _safe_float(admissions.get("Attempts")) or 0.0
    highest_education = "Intermediate" if admissions.get("intermediate") not in (None, "") else "SSC"
    profile_context = {
        "branch": _safe_str(admissions.get("Branch")),
        "batch": _safe_str(admissions.get("Batch")),
        "ssc": admissions.get("ssc"),
        "intermediate": admissions.get("intermediate"),
        "vsat": admissions.get("vsat"),
        "category": _safe_str(admissions.get("Category")),
        "region": _safe_str(admissions.get("Region")),
        "parent_education": _safe_str(admissions.get("ParentEdu")),
        "occupation": _safe_str(admissions.get("Occupation")),
        "income": admissions.get("Income"),
        "registration": {
            "semester": registration.get("Semester"),
            "registered": registration.get("Registered"),
            "final_status": registration.get("FinalStatus"),
        },
        "model_input_exclusions": list(FAIRNESS_REVIEW_REQUIRED_PROFILE_FIELDS),
        "fairness_review_required": True,
        "support_mapping_present": bool(support_mapping_row),
    }

    return {
        "student_id": student_id,
        "student_email": _normalized_email(support_mapping.get("student_email")),
        "faculty_name": _safe_str(support_mapping.get("faculty_name")),
        "faculty_email": _normalized_email(support_mapping.get("faculty_email")),
        "counsellor_name": _safe_str(support_mapping.get("counsellor_name")),
        "counsellor_email": _normalized_email(support_mapping.get("counsellor_email")),
        "parent_name": _safe_str(support_mapping.get("parent_name")),
        "parent_relationship": _normalized_relationship(support_mapping.get("parent_relationship")),
        "parent_email": _normalized_email(support_mapping.get("parent_email")),
        "parent_phone": _normalized_phone(support_mapping.get("parent_phone")),
        "preferred_guardian_channel": _normalized_channel(support_mapping.get("preferred_guardian_channel")),
        "guardian_contact_enabled": _safe_bool(support_mapping.get("guardian_contact_enabled")),
        "external_student_ref": registerno,
        "profile_context": profile_context,
        "gender": gender,
        "highest_education": highest_education,
        "age_band": age_band,
        "disability_status": _normalized_disability_status(support_mapping.get("disability_status")),
        "num_previous_attempts": float(attempts),
    }


def _build_lms_event(student_id: int, registerno: str, row: dict) -> dict:
    event_date = _normalize_event_date(row.get("event_date"))
    return {
        "student_id": student_id,
        "code_module": "VIG",
        "code_presentation": "2026T",
        "id_site": int(_safe_float(row.get("id_site")) or 0),
        "event_date": int(event_date),
        "sum_click": int(_safe_float(row.get("sum_click")) or 0),
        "context_fields": {
            "registerno": registerno,
            "engagement_tag": _safe_str(row.get("engagement_tag")),
            "resource_type": _safe_str(row.get("resource_type")),
            "source_event_id": f"vignan-{registerno}-{event_date}-{row.get('id_site')}",
        },
    }


def _build_erp_event(
    *,
    student_id: int,
    registerno: str,
    academics_rows: list[dict],
    attendance_rows: list[dict],
    admissions_row: dict | None,
    registration_row: dict | None,
) -> dict | None:
    if not academics_rows and not attendance_rows:
        return None

    attendance_ratio = _average_percent(attendance_rows, "Overall%")
    attendance_trend = _average_attendance_trend(attendance_rows, "Trend")
    consecutive_absences = _max_float(attendance_rows, "ConsecutiveAbs")
    missed_sessions = _max_float(attendance_rows, "MissedvDays")
    subject_attendance = {
        _safe_str(row.get("shortname")): _percent_value(row.get("Subject%"))
        for row in attendance_rows
        if _safe_str(row.get("shortname"))
    }
    subject_attendance = {
        key: value for key, value in subject_attendance.items() if value is not None
    }

    avg_marks = _average_float(academics_rows, "Marks")
    avg_cgpa = _average_float(academics_rows, "CGPA")
    backlogs = _max_float(academics_rows, "Backlogs")
    subject_marks = {
        _safe_str(row.get("shortname")): _safe_float(row.get("Marks"))
        for row in academics_rows
        if _safe_str(row.get("shortname"))
    }

    semester = None
    if registration_row:
        semester = registration_row.get("Semester")
    elif attendance_rows:
        semester = attendance_rows[0].get("Semester")
    elif academics_rows:
        semester = academics_rows[0].get("Semester")

    context = {
        "registerno": registerno,
        "attendance_ratio": attendance_ratio,
        "attendance_trend": attendance_trend,
        "consecutive_absences": consecutive_absences,
        "missed_sessions_7d": missed_sessions,
        "subject_attendance": subject_attendance,
        "academic_marks_by_subject": subject_marks,
        "cgpa": avg_cgpa,
        "backlogs": backlogs,
        "backlog_count": backlogs,
        "semester": semester,
        "branch": _safe_str((admissions_row or {}).get("Branch")),
        "source_event_id": f"vignan-{registerno}-erp",
        "event_timestamp": datetime.now().isoformat(),
    }

    weighted_score = avg_marks if avg_marks is not None else (avg_cgpa * 10 if avg_cgpa else 0.0)
    return {
        "student_id": student_id,
        "code_module": "VIG",
        "code_presentation": "2026T",
        "assessment_submission_rate": attendance_ratio,
        "weighted_assessment_score": weighted_score,
        "late_submission_count": int(backlogs or 0),
        "total_assessments_completed": len(academics_rows) if academics_rows else None,
        "assessment_score_trend": attendance_trend,
        "context_fields": context,
    }


def _build_finance_event(student_id: int, registerno: str, row: dict) -> dict:
    status = _safe_str(row.get("Status"))
    return {
        "student_id": student_id,
        "fee_overdue_amount": _safe_float(row.get("Due")),
        "fee_delay_days": int(_safe_float(row.get("DelayDays")) or 0),
        "payment_status": status,
        "modifier_candidate": _safe_float(row.get("Modifier")),
        "context_fields": {
            "registerno": registerno,
            "scholarship": _safe_str(row.get("Scholarship")),
            "status": status,
            "source_event_id": f"vignan-{registerno}-finance",
        },
    }


def _safe_str(value) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()


def _normalized_email(value) -> str | None:
    parsed = _safe_str(value)
    if parsed is None or "@" not in parsed:
        return None
    return parsed.lower()


def _normalized_phone(value) -> str | None:
    parsed = _safe_str(value)
    if parsed is None:
        return None
    cleaned = parsed.replace(" ", "").replace("-", "")
    return cleaned or None


def _normalized_channel(value) -> str | None:
    parsed = _safe_str(value)
    if parsed is None:
        return None
    lowered = parsed.lower()
    if lowered in {"email", "sms", "whatsapp"}:
        return lowered
    return None


def _normalized_relationship(value) -> str | None:
    parsed = _safe_str(value)
    if parsed is None:
        return None
    mapping = {
        "father": "Father",
        "mother": "Mother",
        "parent": "Parent",
        "guardian": "Guardian",
    }
    return mapping.get(parsed.lower(), parsed.title())


def _normalized_disability_status(value) -> str:
    parsed = (_safe_str(value) or "").upper()
    sanitized = parsed.replace("0", "").replace(" ", "")
    if sanitized in {"Y", "YES", "TRUE", "T"}:
        return "Y"
    return "N"


def _safe_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _percent_value(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        text = str(value).strip()
        if text.endswith("%"):
            text = text[:-1]
        return float(text) / 100.0
    except (TypeError, ValueError):
        return None


def _safe_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _average_percent(rows: list[dict], field: str) -> float | None:
    values = [_percent_value(row.get(field)) for row in rows]
    values = [v for v in values if v is not None]
    if not values:
        return None
    return sum(values) / len(values)


def _average_float(rows: list[dict], field: str) -> float | None:
    values = [_safe_float(row.get(field)) for row in rows]
    values = [v for v in values if v is not None]
    if not values:
        return None
    return sum(values) / len(values)


def _max_float(rows: list[dict], field: str) -> float | None:
    values = [_safe_float(row.get(field)) for row in rows]
    values = [v for v in values if v is not None]
    if not values:
        return None
    return max(values)


def _average_attendance_trend(rows: list[dict], field: str) -> float | None:
    values: list[float] = []
    for row in rows:
        raw = row.get(field)
        numeric = _safe_float(raw)
        if numeric is not None:
            values.append(numeric)
            continue
        label = _safe_str(raw)
        if label is None:
            continue
        mapped = {"poor": -1.0, "average": 0.0, "good": 1.0}.get(label.lower())
        if mapped is not None:
            values.append(mapped)
    if not values:
        return None
    return sum(values) / len(values)


def _normalize_event_date(value) -> int:
    if isinstance(value, datetime):
        return value.date().toordinal()
    if isinstance(value, date):
        return value.toordinal()
    if value in (None, ""):
        return int(datetime.now().date().toordinal())
    try:
        return int(float(value))
    except (TypeError, ValueError):
        try:
            parsed = datetime.fromisoformat(str(value))
            return parsed.date().toordinal()
        except ValueError:
            return int(datetime.now().date().toordinal())


def _detect_institution_name(sheets: dict[str, list[dict]]) -> str:
    candidates = (
        sheets.get("AttendancePolicy", [])
        + sheets.get("SubjectCatalog", [])
        + sheets.get("Admissions", [])
    )
    for row in candidates:
        name = _safe_str(row.get("InstitutionName"))
        if name:
            return name
    return "Imported Institution"


def _build_policy_record(row: dict) -> dict:
    institution_name = _safe_str(row.get("InstitutionName")) or "Imported Institution"
    return {
        "institution_name": institution_name,
        "policy_year": _safe_str(row.get("PolicyYear")),
        "overall_min_percent": float(_safe_float(row.get("OverallMinPercent")) or 0.0),
        "subject_min_percent": float(_safe_float(row.get("SubjectMinPercent")) or 0.0),
        "r_grade_below_percent": float(_safe_float(row.get("RGradeBelowPercent")) or 0.0),
        "i_grade_min_percent": float(_safe_float(row.get("IGradeMinPercent")) or 0.0),
        "i_grade_max_percent": float(_safe_float(row.get("IGradeMaxPercent")) or 0.0),
        "condonation_allowed": _safe_bool(row.get("CondonationAllowed")),
        "summer_repeat_for_r": _safe_bool(row.get("SummerRepeatForR")),
        "repeat_internals_for_r": _safe_bool(row.get("RepeatInternalsForR")),
        "end_sem_allowed_for_i": _safe_bool(row.get("EndSemAllowedForI")),
        "end_sem_allowed_for_r": _safe_bool(row.get("EndSemAllowedForR")),
        "policy_notes": _safe_str(row.get("PolicyNotes")),
        "context_fields": {
            "raw_row": {str(key): row.get(key) for key in row.keys()},
        },
    }


def _build_subject_catalog_record(row: dict) -> dict:
    return {
        "institution_name": _safe_str(row.get("InstitutionName")) or "Imported Institution",
        "program_type": _safe_str(row.get("ProgramType")),
        "branch": _safe_str(row.get("Branch")) or "Unknown",
        "regulation": _safe_str(row.get("Regulation")),
        "year": _safe_int(row.get("Year")),
        "semester": _safe_int(row.get("Semester")),
        "subject_code": _safe_str(row.get("SubjectCode")) or "UNKNOWN",
        "subject_name": _safe_str(row.get("SubjectName")) or "Unknown Subject",
        "subject_type": _safe_str(row.get("SubjectType")),
        "credits": _safe_float(row.get("Credits")),
        "is_elective": _safe_bool(row.get("IsElective")),
        "active": not (_safe_str(row.get("Active")) or "yes").strip().lower() in {"no", "false", "0"},
        "context_fields": {
            "raw_row": {str(key): row.get(key) for key in row.keys()},
        },
    }


def _build_subject_catalog_index(records: list[dict]) -> dict[tuple[str, int | None, str], dict]:
    index: dict[tuple[str, int | None, str], dict] = {}
    for record in records:
        branch = str(record.get("branch") or "").strip().lower()
        semester = record.get("semester")
        code = str(record.get("subject_code") or "").strip().lower()
        name = str(record.get("subject_name") or "").strip().lower()
        if code:
            index[(branch, semester, code)] = record
        if name:
            index[(branch, semester, name)] = record
    return index


def _build_profile_payload_generalized(
    *,
    student_id: int,
    registerno: str,
    admissions_row: dict | None,
    registration_row: dict | None,
    support_mapping_row: dict | None,
    academic_progress_row: dict | None,
    institution_name: str,
) -> dict:
    admissions = admissions_row or {}
    registration = registration_row or {}
    support_mapping = support_mapping_row or {}
    academic_progress = academic_progress_row or {}
    semester = _safe_int(registration.get("Semester")) or _safe_int(academic_progress.get("CurrentSemester"))
    current_year = _safe_int(registration.get("CurrentYear")) or _safe_int(academic_progress.get("CurrentYear")) or _year_from_semester(semester)
    gender = _safe_str(admissions.get("Gender")) or "U"
    age_band = _safe_str(admissions.get("AgeBand")) or "Unknown"
    attempts = _safe_float(admissions.get("Attempts")) or 0.0
    highest_education = "Intermediate" if admissions.get("intermediate") not in (None, "") else "SSC"
    profile_context = {
        "institution_name": institution_name,
        "program_type": _safe_str(admissions.get("ProgramType")),
        "branch": _safe_str(admissions.get("Branch")),
        "batch": _safe_str(admissions.get("Batch")),
        "regulation": _safe_str(admissions.get("Regulation")),
        "section": _safe_str(admissions.get("Section")),
        "category": _safe_str(admissions.get("Category")),
        "region": _safe_str(admissions.get("Region")),
        "parent_education": _safe_str(admissions.get("ParentEdu")),
        "occupation": _safe_str(admissions.get("Occupation")),
        "income": admissions.get("Income"),
        "registration": {
            "semester": semester,
            "current_year": current_year,
            "registered": registration.get("Registered"),
            "final_status": registration.get("FinalStatus"),
            "academic_year": registration.get("AcademicYear"),
            "semester_mode": _safe_str(registration.get("SemesterMode")) or _safe_str(academic_progress.get("SemesterMode")),
        },
        "support_contacts": {
            "student_phone": _normalized_phone(support_mapping.get("student_phone")),
            "faculty_phone": _normalized_phone(support_mapping.get("faculty_phone")),
            "counsellor_phone": _normalized_phone(support_mapping.get("counsellor_phone")),
        },
        "student_academic_progress": {
            "current_status": _safe_str(academic_progress.get("CurrentAcademicStatus")),
            "standing_label": _safe_str(academic_progress.get("StandingLabel")),
            "expected_graduation_year": _safe_int(academic_progress.get("ExpectedGraduationYear")),
            "total_backlogs": _safe_int(academic_progress.get("TotalBacklogs")),
        },
        "model_input_exclusions": list(FAIRNESS_REVIEW_REQUIRED_PROFILE_FIELDS),
        "fairness_review_required": True,
        "support_mapping_present": bool(support_mapping_row),
    }
    return {
        "student_id": student_id,
        "student_email": _normalized_email(support_mapping.get("student_email")),
        "faculty_name": _safe_str(support_mapping.get("faculty_name")),
        "faculty_email": _normalized_email(support_mapping.get("faculty_email")),
        "counsellor_name": _safe_str(support_mapping.get("counsellor_name")),
        "counsellor_email": _normalized_email(support_mapping.get("counsellor_email")),
        "parent_name": _safe_str(support_mapping.get("parent_name")),
        "parent_relationship": _normalized_relationship(support_mapping.get("parent_relationship")),
        "parent_email": _normalized_email(support_mapping.get("parent_email")),
        "parent_phone": _normalized_phone(support_mapping.get("parent_phone")),
        "preferred_guardian_channel": _normalized_channel(support_mapping.get("preferred_guardian_channel")),
        "guardian_contact_enabled": _safe_bool(support_mapping.get("guardian_contact_enabled")),
        "external_student_ref": registerno,
        "profile_context": profile_context,
        "gender": gender,
        "highest_education": highest_education,
        "age_band": age_band,
        "disability_status": _normalized_disability_status(support_mapping.get("disability_status")),
        "num_previous_attempts": float(attempts),
    }


def _build_student_academic_progress_record(
    *,
    student_id: int,
    registerno: str,
    institution_name: str,
    admissions_row: dict | None,
    registration_row: dict | None,
    academic_progress_row: dict | None,
) -> dict | None:
    admissions = admissions_row or {}
    registration = registration_row or {}
    progress = academic_progress_row or {}
    semester = _safe_int(progress.get("CurrentSemester")) or _safe_int(registration.get("Semester"))
    current_year = _safe_int(progress.get("CurrentYear")) or _safe_int(registration.get("CurrentYear")) or _year_from_semester(semester)
    if semester is None and current_year is None and not admissions and not progress:
        return None
    return {
        "student_id": student_id,
        "external_student_ref": registerno,
        "institution_name": institution_name,
        "program_type": _safe_str(admissions.get("ProgramType")),
        "branch": _safe_str(admissions.get("Branch")),
        "batch": _safe_str(admissions.get("Batch")),
        "current_year": current_year,
        "current_semester": semester,
        "current_academic_status": _safe_str(progress.get("CurrentAcademicStatus")) or _safe_str(registration.get("FinalStatus")),
        "semester_mode": _safe_str(progress.get("SemesterMode")) or _safe_str(registration.get("SemesterMode")) or "regular_coursework",
        "expected_graduation_year": _safe_int(progress.get("ExpectedGraduationYear")),
        "standing_label": _safe_str(progress.get("StandingLabel")),
        "total_backlogs": _safe_int(progress.get("TotalBacklogs")),
        "context_fields": {
            "raw_progress_row": {str(key): progress.get(key) for key in progress.keys()},
            "registration_final_status": _safe_str(registration.get("FinalStatus")),
        },
    }


def _build_subject_attendance_records(
    *,
    student_id: int,
    registerno: str,
    institution_name: str,
    admissions_row: dict | None,
    attendance_rows: list[dict],
    policy: dict,
    subject_catalog_index: dict[tuple[str, int | None, str], dict],
) -> list[dict]:
    admissions = admissions_row or {}
    branch = _safe_str(admissions.get("Branch")) or "Unknown"
    built: list[dict] = []
    for row in attendance_rows:
        semester = _safe_int(row.get("Semester"))
        overall_percent = _normalize_percent_number(
            _row_value(row, "Overall%", "OverallPer", "OverallPercent", "OverallAttendancePercent")
        )
        subject_percent = _normalize_percent_number(
            _row_value(row, "Subject%", "SubjectPer", "SubjectPercent", "SubjectAttendancePercent")
        )
        subject_identifier = (
            _safe_str(_row_value(row, "SubjectCode", "SubjectCo"))
            or _safe_str(_row_value(row, "shortname", "SubjectName", "Subject"))
            or "Unknown Subject"
        )
        subject_catalog = _lookup_subject_catalog(
            subject_catalog_index,
            branch=branch,
            semester=semester,
            subject_identifier=subject_identifier,
        )
        policy_result = _evaluate_attendance_policy(
            overall_percent=overall_percent,
            subject_percent=subject_percent,
            policy=policy,
        )
        subject_name = (
            _safe_str(_row_value(row, "SubjectName", "shortname", "Subject"))
            or (subject_catalog or {}).get("subject_name")
            or _safe_str(_row_value(row, "shortname", "Subject"))
            or subject_identifier
        )
        built.append(
            {
                "student_id": student_id,
                "external_student_ref": registerno,
                "institution_name": institution_name,
                "branch": branch,
                "year": _safe_int(row.get("Year")) or _year_from_semester(semester),
                "semester": semester,
                "subject_code": _safe_str(_row_value(row, "SubjectCode", "SubjectCo")) or (subject_catalog or {}).get("subject_code"),
                "subject_name": subject_name,
                "subject_type": _safe_str(row.get("SubjectType")) or (subject_catalog or {}).get("subject_type"),
                "overall_attendance_percent": overall_percent,
                "subject_attendance_percent": subject_percent,
                "required_percent": float(policy.get("subject_min_percent") or 0.0),
                "overall_status": policy_result["overall_status"],
                "subject_status": policy_result["subject_status"],
                "grade_consequence": policy_result["grade_consequence"],
                "condonation_required": policy_result["condonation_required"],
                "summer_repeat_required": policy_result["summer_repeat_required"],
                "internals_repeat_required": policy_result["internals_repeat_required"],
                "end_sem_eligible": policy_result["end_sem_eligible"],
                "classes_conducted": _safe_int(row.get("ClassesConducted")),
                "classes_attended": _safe_int(row.get("ClassesAttended")),
                "consecutive_absences": _safe_int(row.get("ConsecutiveAbs")),
                "missed_days": _safe_int(row.get("MissedvDays")),
                "trend": _normalized_trend_label(row.get("Trend")),
                "context_fields": {
                    "catalog_match": bool(subject_catalog),
                    "raw_row": {str(key): row.get(key) for key in row.keys()},
                },
            }
        )
    return built


def _build_semester_progress_records(
    *,
    student_id: int,
    registerno: str,
    provided_rows: list[dict],
    attendance_rows: list[dict],
    registration_row: dict | None,
    academic_progress_row: dict | None,
    policy: dict,
) -> list[dict]:
    if provided_rows:
        built: list[dict] = []
        for row in provided_rows:
            built.append(
                {
                    "student_id": student_id,
                    "external_student_ref": registerno,
                    "year": _safe_int(row.get("Year")) or _year_from_semester(_safe_int(row.get("Semester"))),
                    "semester": _safe_int(row.get("Semester")),
                    "overall_attendance_percent": _normalize_percent_number(row.get("OverallAttendancePercent")),
                    "overall_status": _safe_str(row.get("OverallStatus")),
                    "subjects_below_75_count": _safe_int(row.get("SubjectsBelow75")),
                    "subjects_below_65_count": _safe_int(row.get("SubjectsBelow65")),
                    "has_i_grade_risk": _safe_bool(row.get("HasIGradeRisk")),
                    "has_r_grade_risk": _safe_bool(row.get("HasRGradeRisk")),
                    "current_eligibility": _safe_str(row.get("CurrentEligibility")),
                    "semester_mode": _safe_str(row.get("SemesterMode")),
                    "context_fields": {"raw_row": {str(key): row.get(key) for key in row.keys()}},
                }
            )
        return built

    grouped: dict[tuple[int | None, int | None], list[dict]] = defaultdict(list)
    for row in attendance_rows:
        grouped[(row.get("year"), row.get("semester"))].append(row)
    built = []
    current_mode = _safe_str((academic_progress_row or {}).get("SemesterMode")) or _safe_str((registration_row or {}).get("SemesterMode"))
    for (year, semester), rows in sorted(grouped.items(), key=lambda item: ((item[0][0] or 0), (item[0][1] or 0))):
        subject_statuses = [str(row.get("subject_status") or "").upper() for row in rows]
        overall_values = [row.get("overall_attendance_percent") for row in rows if row.get("overall_attendance_percent") is not None]
        overall_percent = (sum(overall_values) / len(overall_values)) if overall_values else None
        built.append(
            {
                "student_id": student_id,
                "external_student_ref": registerno,
                "year": year,
                "semester": semester,
                "overall_attendance_percent": overall_percent,
                "overall_status": "SAFE" if (overall_percent or 0.0) >= float(policy.get("overall_min_percent") or 0.0) else "SHORTAGE",
                "subjects_below_75_count": sum(1 for status_value in subject_statuses if status_value in {"I_GRADE", "R_GRADE"}),
                "subjects_below_65_count": sum(1 for status_value in subject_statuses if status_value == "R_GRADE"),
                "has_i_grade_risk": any(status_value == "I_GRADE" for status_value in subject_statuses),
                "has_r_grade_risk": any(status_value == "R_GRADE" for status_value in subject_statuses),
                "current_eligibility": "Eligible" if all(bool(row.get("end_sem_eligible")) for row in rows) else "At Risk",
                "semester_mode": current_mode or "regular_coursework",
                "context_fields": {"derived_from_attendance": True},
            }
        )
    return built


def _build_student_academic_records(
    *,
    student_id: int,
    registerno: str,
    institution_name: str,
    admissions_row: dict | None,
    source_rows: list[dict],
    subject_catalog_index: dict[tuple[str, int | None, str], dict],
    attendance_rows: list[dict],
) -> list[dict]:
    admissions = admissions_row or {}
    branch = _safe_str(admissions.get("Branch")) or "Unknown"
    attendance_index: dict[tuple[int | None, str], dict] = {}
    for row in attendance_rows:
        subject_key = _safe_str(row.get("subject_code")) or _safe_str(row.get("subject_name")) or ""
        attendance_index[(row.get("semester"), subject_key.lower())] = row
    built: list[dict] = []
    for row in source_rows:
        semester = _safe_int(row.get("Semester"))
        subject_identifier = _safe_str(row.get("SubjectCode")) or _safe_str(row.get("shortname")) or _safe_str(row.get("SubjectName")) or "Unknown Subject"
        subject_catalog = _lookup_subject_catalog(
            subject_catalog_index,
            branch=branch,
            semester=semester,
            subject_identifier=subject_identifier,
        )
        subject_code = _safe_str(row.get("SubjectCode")) or (subject_catalog or {}).get("subject_code")
        subject_name = _safe_str(row.get("SubjectName")) or (subject_catalog or {}).get("subject_name") or _safe_str(row.get("shortname")) or subject_identifier
        attendance_match = attendance_index.get((semester, (subject_code or subject_name).lower()))
        raw_grade = _safe_str(row.get("Grade"))
        raw_result_status = _safe_str(row.get("ResultStatus"))
        attendance_linked_status = _safe_str(row.get("AttendanceLinkedStatus")) or _safe_str((attendance_match or {}).get("subject_status"))
        effective_grade, effective_result_status, overridden = _effective_academic_outcome(
            raw_grade=raw_grade,
            raw_result_status=raw_result_status,
            attendance_linked_status=attendance_linked_status,
        )
        built.append(
            {
                "student_id": student_id,
                "external_student_ref": registerno,
                "institution_name": institution_name,
                "branch": branch,
                "year": _safe_int(row.get("Year")) or _year_from_semester(semester),
                "semester": semester,
                "subject_code": subject_code,
                "subject_name": subject_name,
                "credits": _safe_float(row.get("Credits")) or (subject_catalog or {}).get("credits"),
                "internal_marks": _safe_float(row.get("InternalMarks")),
                "external_marks": _safe_float(row.get("ExternalMarks")),
                "total_marks": _safe_float(row.get("TotalMarks")),
                "marks": _safe_float(row.get("Marks")),
                "grade": effective_grade,
                "result_status": effective_result_status,
                "attendance_linked_status": attendance_linked_status,
                "cgpa": _safe_float(row.get("CGPA")),
                "backlogs": _safe_float(row.get("Backlogs")),
                "context_fields": {
                    "raw_row": {str(key): row.get(key) for key in row.keys()},
                    "catalog_match": bool(subject_catalog),
                    "raw_grade": raw_grade,
                    "raw_result_status": raw_result_status,
                    "effective_outcome_overridden": overridden,
                },
            }
        )
    return built


def _effective_academic_outcome(
    *,
    raw_grade: str | None,
    raw_result_status: str | None,
    attendance_linked_status: str | None,
) -> tuple[str | None, str | None, bool]:
    normalized_status = _normalize_uncleared_token(attendance_linked_status)
    normalized_grade = _normalize_uncleared_token(raw_grade)
    normalized_result = _normalize_uncleared_token(raw_result_status)

    effective_marker = normalized_status or normalized_grade or normalized_result
    if effective_marker == "R_GRADE":
        return "R", "Pending R-grade clearance", True
    if effective_marker == "I_GRADE":
        return "I", "Pending I-grade clearance", True
    return raw_grade, raw_result_status, False


def _normalize_uncleared_token(value: str | None) -> str | None:
    parsed = _safe_str(value)
    if parsed is None:
        return None
    normalized = parsed.strip().upper().replace("-", "_").replace(" ", "_")
    if normalized in {"R", "R_GRADE", "REPEAT", "REPEAT_GRADE"}:
        return "R_GRADE"
    if normalized in {"I", "I_GRADE", "INCOMPLETE"}:
        return "I_GRADE"
    return None


def _build_lms_event_generic(student_id: int, registerno: str, row: dict) -> dict:
    event_date = _normalize_event_date(row.get("event_date"))
    institution_code = _safe_str(row.get("InstitutionCode")) or "INST"
    presentation = _safe_str(row.get("AcademicYear")) or "CURRENT"
    return {
        "student_id": student_id,
        "code_module": institution_code[:20],
        "code_presentation": presentation[:20],
        "id_site": int(_safe_float(row.get("id_site")) or 0),
        "event_date": int(event_date),
        "sum_click": int(_safe_float(row.get("sum_click")) or 0),
        "context_fields": {
            "registerno": registerno,
            "engagement_tag": _safe_str(row.get("engagement_tag")),
            "resource_type": _safe_str(row.get("resource_type")),
            "subject_code": _safe_str(row.get("SubjectCode")),
            "subject_name": _safe_str(row.get("SubjectName")),
            "source_event_id": f"institution-{registerno}-{event_date}-{row.get('id_site')}",
        },
    }


def _build_finance_event_generic(student_id: int, registerno: str, row: dict) -> dict:
    status = _safe_str(row.get("Status"))
    return {
        "student_id": student_id,
        "fee_overdue_amount": _safe_float(row.get("Due")),
        "fee_delay_days": int(_safe_float(row.get("DelayDays")) or 0),
        "payment_status": status,
        "modifier_candidate": _safe_float(row.get("Modifier")),
        "context_fields": {
            "registerno": registerno,
            "scholarship": _safe_str(row.get("Scholarship")),
            "status": status,
            "fee_type": _safe_str(row.get("FeeType")),
            "installment_status": _safe_str(row.get("InstallmentStatus")),
            "source_event_id": f"institution-{registerno}-finance",
        },
    }


def _build_erp_event_generalized(
    *,
    student_id: int,
    registerno: str,
    institution_name: str,
    admissions_row: dict | None,
    registration_row: dict | None,
    attendance_rows: list[dict],
    academic_rows: list[dict],
) -> dict | None:
    if not attendance_rows and not academic_rows:
        return None
    attendance_values = [row.get("overall_attendance_percent") for row in attendance_rows if row.get("overall_attendance_percent") is not None]
    attendance_ratio = ((sum(attendance_values) / len(attendance_values)) / 100.0) if attendance_values else None
    subject_attendance = {
        str(row.get("subject_name")): (row.get("subject_attendance_percent") / 100.0 if row.get("subject_attendance_percent") is not None else None)
        for row in attendance_rows
        if row.get("subject_name")
    }
    subject_attendance = {key: value for key, value in subject_attendance.items() if value is not None}
    marks_values = [row.get("marks") for row in academic_rows if row.get("marks") is not None]
    cgpa_values = [row.get("cgpa") for row in academic_rows if row.get("cgpa") is not None]
    backlog_values = [row.get("backlogs") for row in academic_rows if row.get("backlogs") is not None]
    weighted_score = (sum(marks_values) / len(marks_values)) if marks_values else ((sum(cgpa_values) / len(cgpa_values)) * 10 if cgpa_values else 0.0)
    latest_semester = _safe_int((registration_row or {}).get("Semester")) or max((row.get("semester") for row in attendance_rows if row.get("semester") is not None), default=None)
    return {
        "student_id": student_id,
        "code_module": (_safe_str((admissions_row or {}).get("Branch")) or institution_name or "INST")[:20],
        "code_presentation": (_safe_str((registration_row or {}).get("AcademicYear")) or "CURRENT")[:20],
        "assessment_submission_rate": attendance_ratio,
        "weighted_assessment_score": weighted_score,
        "late_submission_count": int(max(backlog_values) if backlog_values else 0),
        "total_assessments_completed": len(academic_rows) if academic_rows else None,
        "assessment_score_trend": _average_float(
            [{"value": _trend_to_numeric(row.get("trend"))} for row in attendance_rows if row.get("trend") is not None],
            "value",
        ),
        "context_fields": {
            "registerno": registerno,
            "institution_name": institution_name,
            "semester": latest_semester,
            "branch": _safe_str((admissions_row or {}).get("Branch")),
            "attendance_ratio": attendance_ratio,
            "subject_attendance": subject_attendance,
            "low_attendance_subjects": [
                str(row.get("subject_name"))
                for row in attendance_rows
                if str(row.get("subject_status") or "").upper() in {"I_GRADE", "R_GRADE"}
            ],
            "subject_grade_statuses": {
                str(row.get("subject_name")): row.get("subject_status")
                for row in attendance_rows
                if row.get("subject_name")
            },
            "source_event_id": f"institution-{registerno}-erp",
            "event_timestamp": datetime.now().isoformat(),
        },
    }


def _lookup_subject_catalog(
    subject_catalog_index: dict[tuple[str, int | None, str], dict],
    *,
    branch: str,
    semester: int | None,
    subject_identifier: str,
) -> dict | None:
    branch_key = str(branch or "").strip().lower()
    identifier_key = str(subject_identifier or "").strip().lower()
    if not identifier_key:
        return None
    return subject_catalog_index.get((branch_key, semester, identifier_key))


def _evaluate_attendance_policy(
    *,
    overall_percent: float | None,
    subject_percent: float | None,
    policy: dict,
) -> dict:
    overall_min = float(policy.get("overall_min_percent") or 0.0)
    subject_min = float(policy.get("subject_min_percent") or 0.0)
    r_grade_below = float(policy.get("r_grade_below_percent") or 0.0)
    i_grade_min = float(policy.get("i_grade_min_percent") or 0.0)
    i_grade_max = float(policy.get("i_grade_max_percent") or 0.0)

    overall_status = "SAFE" if (overall_percent or 0.0) >= overall_min else "SHORTAGE"
    subject_status = "SAFE"
    grade_consequence = "Safe"
    condonation_required = False
    summer_repeat_required = False
    internals_repeat_required = False
    end_sem_eligible = True

    if subject_percent is None:
        subject_status = "UNAVAILABLE"
        grade_consequence = "Attendance data unavailable"
        end_sem_eligible = overall_status == "SAFE"
    elif subject_percent < r_grade_below:
        subject_status = "R_GRADE"
        grade_consequence = "Repeat subject in summer with internals and end-sem"
        summer_repeat_required = bool(policy.get("summer_repeat_for_r"))
        internals_repeat_required = bool(policy.get("repeat_internals_for_r"))
        end_sem_eligible = bool(policy.get("end_sem_allowed_for_r"))
    elif i_grade_min <= subject_percent <= i_grade_max and subject_percent < subject_min:
        subject_status = "I_GRADE"
        grade_consequence = "Condonation required before end-sem"
        condonation_required = bool(policy.get("condonation_allowed"))
        end_sem_eligible = bool(policy.get("end_sem_allowed_for_i"))
    elif subject_percent < subject_min:
        subject_status = "SHORTAGE"
        grade_consequence = "Attendance shortage requires institutional review"
        end_sem_eligible = False

    return {
        "overall_status": overall_status,
        "subject_status": subject_status,
        "grade_consequence": grade_consequence,
        "condonation_required": condonation_required,
        "summer_repeat_required": summer_repeat_required,
        "internals_repeat_required": internals_repeat_required,
        "end_sem_eligible": end_sem_eligible and overall_status == "SAFE",
    }


def _normalize_percent_number(value) -> float | None:
    parsed = _percent_value(value)
    if parsed is not None:
        return round(parsed * 100.0, 2)
    raw = _safe_float(value)
    if raw is None:
        return None
    if raw <= 1.0:
        return round(raw * 100.0, 2)
    return round(raw, 2)


def _year_from_semester(semester: int | None) -> int | None:
    if semester is None:
        return None
    return int((semester + 1) / 2)


def _safe_int(value) -> int | None:
    parsed = _safe_float(value)
    if parsed is None:
        return None
    return int(parsed)


def _normalized_trend_label(value) -> str | None:
    parsed = _safe_str(value)
    if parsed is None:
        return None
    numeric = _safe_float(parsed)
    if numeric is not None:
        if numeric > 0:
            return "Improving"
        if numeric < 0:
            return "Declining"
        return "Stable"
    lowered = parsed.lower()
    if lowered in {"good", "improving", "up"}:
        return "Improving"
    if lowered in {"poor", "declining", "down"}:
        return "Declining"
    return "Stable"


def _trend_to_numeric(value) -> float | None:
    label = _normalized_trend_label(value)
    if label == "Improving":
        return 1.0
    if label == "Declining":
        return -1.0
    if label == "Stable":
        return 0.0
    return None
